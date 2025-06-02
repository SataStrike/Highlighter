"""
Excel Helper Module - Handles Excel file operations with formatting
"""
import pandas as pd
import openpyxl
import os
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Global debug flag
DEBUG = True

# Global configuration for formatting
FORMATTING_CONFIG = {
    'apply_to_entire_row': False,  # Whether to format entire row or just domain cell
    'format_summary_sheet': True,  # Whether to format the summary sheet
    'format_supply_chain': False,  # Whether to format the supply chain sheet
    'custom_styles': {
        # Can be overridden if needed
        "High": {
            "bg_color": "FFC7CE",  # Light red
            "font_color": "9C0006"  # Dark red
        },
        "Medium": {
            "bg_color": "FFEB9C",  # Light yellow
            "font_color": "9C6500"  # Dark yellow
        },
        "Low": {
            "bg_color": "C6EFCE",  # Light green
            "font_color": "006100"  # Dark green
        }
    }
}

# No need for global variables to track website relationships

def debug_print(*args, **kwargs):
    """Print debug information only if DEBUG is True"""
    if DEBUG:
        print("DEBUG:", *args, **kwargs)

def evaluate_rule(cell_value, operator, value):
    """
    Evaluate if a cell value satisfies a rule.
    
    Parameters:
        cell_value: The value in the cell
        operator (str): One of '>', '<', '=', 'Between'
        value: The threshold value or range to compare against
        
    Returns:
        bool: True if rule is satisfied, False otherwise
    """
    try:
        # Skip if cell_value is missing
        if cell_value is None or pd.isna(cell_value):
            return False
            
        # Print debugging for the green highlight rule (> 5)
        if operator == '>' and str(value) == '5':
            print(f"DEBUG: Green highlight rule check: {cell_value} > 5")
            
        # Convert to numeric if possible
        try:
            if isinstance(cell_value, str):
                # Remove % sign if present in the string
                if '%' in cell_value:
                    cell_value = cell_value.replace('%', '')
            cell_value = float(cell_value)
        except (ValueError, TypeError):
            # If not numeric, treat as string
            pass
            
        # Handle different operators
        if operator == '>':
            try:
                rule_value = float(value)
                # Convert percentage values (if they're decimals like 0.xx) to actual percentages
                if isinstance(cell_value, float) and cell_value < 1 and 'Rate' in str(operator):
                    cell_value = cell_value * 100
                result = cell_value > rule_value
                if str(value) == '5':
                    print(f"DEBUG: {cell_value} > {rule_value} = {result}")
                return result
            except (ValueError, TypeError) as e:
                print(f"Error comparing {cell_value} > {value}: {str(e)}")
                return False
        elif operator == '<':
            rule_value = float(value)
            return cell_value < rule_value
        elif operator == '=':
            # For string comparison
            if isinstance(cell_value, str):
                return cell_value.lower() == str(value).lower()
            else:
                rule_value = float(value)
                return cell_value == rule_value
        elif operator == 'Between':
            # Parse range (format: "min;max")
            if ';' in str(value):
                parts = str(value).split(';')
                if len(parts) == 2:
                    try:
                        min_val = float(parts[0])
                        max_val = float(parts[1])
                        return min_val <= cell_value <= max_val
                    except (ValueError, TypeError):
                        return False
                        
        return False
    except Exception as e:
        print(f"Error evaluating rule: {str(e)}")
        return False

def format_domains_highlight_sheet(excel_path):
    """
    Apply specialized formatting to the Domains Highlight sheet:
    1. Limit all decimal values to two decimal places
    2. Format all difference columns as percentages
    3. Format Bid Rate, Win Rate, Fill Rate, and Viewability as percentages
    4. Apply conditional formatting to difference columns (green for positive, red for negative)
    
    Parameters:
        excel_path (str): Path to the Excel file to modify
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)
        
        # Check if Domains Highlight sheet exists
        if 'Domains Highlight' not in workbook.sheetnames:
            print("Domains Highlight sheet not found, skipping custom formatting")
            return
            
        sheet = workbook['Domains Highlight']
        
        # Get column indices for special formatting
        rate_columns = []  # For columns to format as percentages
        diff_columns = []  # For difference columns
        decimal_columns = []  # For columns to format with 2 decimal places
        
        for col_idx, cell in enumerate(sheet[1], 1):
            col_name = cell.value
            if not col_name:
                continue
                
            col_name_lower = str(col_name).lower()
            
            # Find rate columns to format as percentages
            if any(rate_term in col_name_lower for rate_term in ['bid rate', 'win rate', 'fill rate', 'viewability']):
                rate_columns.append(col_idx)
                
            # Find difference columns for conditional formatting
            if 'diff' in col_name_lower or 'difference' in col_name_lower:
                diff_columns.append(col_idx)
                
            # Any numeric column should have 2 decimal places
            if any(term in col_name_lower for term in ['rate', 'revenue', 'rpm', 'cpm', 'ecpm', 'rpb', 'value', 'cost', 'price', '%', 'time in view']):
                decimal_columns.append(col_idx)
        
        # Add rate columns to decimal columns list if not already there
        for col_idx in rate_columns:
            if col_idx not in decimal_columns:
                decimal_columns.append(col_idx)
        
        # Add diff columns to decimal columns list if not already there
        for col_idx in diff_columns:
            if col_idx not in decimal_columns:
                decimal_columns.append(col_idx)
        
        print(f"Formatting columns in Domains Highlight sheet:")
        print(f"  - Rate columns (as percentages): {rate_columns}")
        print(f"  - Diff columns (conditional format): {diff_columns}")
        print(f"  - Decimal columns (2 decimal places): {decimal_columns}")
        
        # Apply formatting to each row
        for row_idx in range(2, sheet.max_row + 1):
            # Format decimal columns to 2 decimal places
            for col_idx in decimal_columns:
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
            
            # Format rate columns as percentages
            for col_idx in rate_columns:
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'
                    
            # Format diff columns as percentages with conditional formatting
            for col_idx in diff_columns:
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    # Excel multiplies by 100 when applying % format, so divide by 100 first
                    # Store original value to check if positive/negative
                    original_value = cell.value
                    
                    # Divide by 100 before applying percentage format
                    # This way -69.71 will display as -69.71% rather than -6971%
                    cell.value = cell.value / 100
                    
                    # Format as percentage
                    cell.number_format = '0.00%'
                    
                    # Apply conditional formatting based on positive/negative of original value
                    if original_value > 0:
                        # Green text for positive values
                        cell.font = Font(color='006100')  # Dark green
                    elif original_value < 0:
                        # Red text for negative values
                        cell.font = Font(color='9C0006')  # Dark red
        
        # Save the workbook
        workbook.save(excel_path)
        print("Applied custom formatting to Domains Highlight sheet")
        
    except Exception as e:
        print(f"Error applying custom formatting: {str(e)}")
        import traceback
        traceback.print_exc()

def write_to_excel_with_two_sheets(output_path, domains_df, supply_chain_df, apply_formatting=True, rules=None):
    """
    Write two dataframes to an Excel file with each dataframe on its own sheet.
    Now supports two-phase processing with separate formatting.
    
    Parameters:
        output_path (str): Path to the Excel file to write to
        domains_df (DataFrame): DataFrame containing domains highlight data
        supply_chain_df (DataFrame): DataFrame containing supply chain validation data
        apply_formatting (bool): Whether to apply formatting in this call or defer it
        rules (dict): Dictionary of rules to apply for formatting
        
    Returns:
        bool: True if successful, False otherwise
    """
    success = False
    try:
        # Phase 1: Data Population Only
        print(f"Phase 1: Writing data to {output_path}")
        
        # Check if the file already exists (it should, created by the domains_highlighter)
        if os.path.exists(output_path):
            print(f"DEBUG: File exists, loading workbook")
            # Load the existing workbook
            book = openpyxl.load_workbook(output_path)
            
            # Check if the Supply Chain Validation sheet already exists
            sheet_name = 'Supply Chain Validation'
            if sheet_name in book.sheetnames:
                print(f"DEBUG: {sheet_name} already exists, removing it first")
                # Remove the sheet if it exists
                idx = book.sheetnames.index(sheet_name)
                book.remove(book.worksheets[idx])
            
            # Save the workbook
            book.save(output_path)
            
            # Now create a writer and add the Supply Chain sheet
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                print(f"DEBUG: Writing Supply Chain data to sheet")
                supply_chain_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
        else:
            # If for some reason the file doesn't exist, create it from scratch
            print(f"DEBUG: File does not exist, creating from scratch")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                domains_df.to_excel(writer, sheet_name='Domains Highlight', index=False)
                supply_chain_df.to_excel(writer, sheet_name='Supply Chain Validation', index=False)
        
        print(f"DEBUG: Successfully added Supply Chain sheet to {output_path}")
        
        # Now create a summary sheet with data from both sheets (no formatting yet)
        try:
            print(f"Creating summary sheet (data only)...")
            create_summary_sheet(output_path, domains_df, supply_chain_df, apply_formatting=False)
            print(f"DEBUG: Successfully added Summary sheet to {output_path}")
        except Exception as e:
            print(f"Error creating summary sheet: {str(e)}")
            import traceback
            traceback.print_exc()
            # Continue even if summary sheet fails
        
        # Apply custom formatting to the Domains Highlight sheet
        format_domains_highlight_sheet(output_path)
        
        # Phase 2: Apply formatting if requested
        if apply_formatting and rules is not None:
            print("Phase 2: Applying consistent formatting to all sheets")
            apply_consistent_formatting(output_path, domains_df, rules)
            
        success = True
        return success
    except Exception as e:
        print(f"Error writing Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def apply_custom_domain_styling_to_excel(
    excel_path,
    sheet_name,
    column_name,
    condition_fn,
    bg_color='#FFFF00',
    font_color='#000000',
):
    """
    Apply custom conditional formatting to a column in an Excel file.
    
    Parameters:
        excel_path (str): Path to the Excel file to modify.
        sheet_name (str): Name of the worksheet to apply formatting.
        column_name (str): Name of the column to apply formatting.
        condition_fn (callable): Function that takes a cell value and returns True if formatting should be applied.
        bg_color (str): Background color in hex (e.g., 'FFFF00' for yellow).
        font_color (str): Font color in hex (e.g., '000000' for black).
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    # Find the column index for the specified column name
    header = [cell.value for cell in ws[1]]
    try:
        col_idx = header.index(column_name) + 1  # 1-based index
    except ValueError:
        raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}'")

    fill = PatternFill(start_color=bg_color.replace('#',''), end_color=bg_color.replace('#',''), fill_type='solid')
    font = Font(color=font_color.replace('#',''))

    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if condition_fn(cell.value):
            cell.fill = fill
            cell.font = font

    wb.save(excel_path)

def add_formatting_legends(sheet, legend_title, legend_items, start_cell=None, include_descriptions=True):
    """
    Add a legend to explain conditional formatting.
    
    Parameters:
        sheet: The worksheet to add the legend to
        legend_title: The title of the legend
        legend_items: List of dicts with keys 'label', 'bg_color', and 'font_color'
        start_cell: Optional tuple (row, column) to specify where to start the legend
                    If None, will use (sheet.max_row + 3, 1)
        include_descriptions: Whether to include description cells
    
    Returns:
        tuple: The next position (row, column) after the legend
    """
    # Determine start position
    if start_cell is None:
        start_row = sheet.max_row + 3
        start_col = 1
    else:
        start_row, start_col = start_cell
    
    # Add legend title
    title_cell = sheet.cell(row=start_row, column=start_col)
    title_cell.value = legend_title
    title_cell.font = Font(bold=True)
    
    # Add legend items
    for i, item in enumerate(legend_items):
        row = start_row + i + 1
        
        # Sample cell with formatting
        sample_cell = sheet.cell(row=row, column=start_col)
        sample_cell.value = item['label']
        sample_cell.fill = PatternFill(start_color=item['bg_color'], end_color=item['bg_color'], fill_type='solid')
        sample_cell.font = Font(color=item['font_color'])
        
        # Description cell (if needed)
        if include_descriptions and 'description' in item:
            desc_cell = sheet.cell(row=row, column=start_col + 1)
            desc_cell.value = item['description']
            # Merge cells if not at the edge of the sheet
            if start_col + 4 <= sheet.max_column:
                sheet.merge_cells(start_row=row, start_column=start_col + 1, 
                                 end_row=row, end_column=start_col + 4)
    
    # Return the next position after the legend
    return (start_row + len(legend_items) + 2, start_col)

def write_error_distribution_sheet(output_path, error_dist_df, error_dist_summary, apply_formatting=True):
    """
    Write error distribution data to a new sheet in the Excel file and update the Summary sheet
    with error distribution summary columns (Most Adcalls status and Most Ad Calls %)
    
    Parameters:
        output_path (str): Path to the Excel file
        error_dist_df (DataFrame): DataFrame with error distribution data
        error_dist_summary (dict): Dictionary with summary data for each website
        apply_formatting (bool): Whether to apply formatting immediately
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        debug_print(f"Writing error distribution data to {output_path}")
        
        # Load the workbook
        book = openpyxl.load_workbook(output_path)
        
        # Check if Error Distribution sheet already exists
        sheet_name = 'Error Distribution'
        if sheet_name in book.sheetnames:
            debug_print(f"{sheet_name} already exists, removing it first")
            idx = book.sheetnames.index(sheet_name)
            book.remove(book.worksheets[idx])
        
        # Save the workbook
        book.save(output_path)
        
        # Write the error distribution data
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            error_dist_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply basic formatting to the error distribution sheet
        book = openpyxl.load_workbook(output_path)
        sheet = book[sheet_name]
        
        # Format headers
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(bottom=Side(style='medium', color='000000'))
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            column_header = sheet.cell(row=1, column=column[0].column).value
            
            # For other columns, calculate based on content
            for cell in column:
                if cell.row == 1:  # Header row
                    max_length = max(max_length, len(str(cell.value)) + 2)  # Extra padding for headers
                else:
                    max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
            
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add filters
        sheet.auto_filter.ref = sheet.dimensions
        
        # Update the Summary sheet with error distribution summary columns if it exists
        if 'Summary' in book.sheetnames and error_dist_summary:
            summary_sheet = book['Summary']
            
            # Find the Website/App Name column in the Summary sheet
            website_col_idx = None
            for col_idx, cell in enumerate(summary_sheet[1], 1):
                if cell.value == 'Website/App Name':
                    website_col_idx = col_idx
                    break
            
            if website_col_idx is not None:
                # Add new column headers for error distribution summary
                status_col_idx = summary_sheet.max_column + 1
                pct_col_idx = status_col_idx + 1
                
                # Add column headers
                status_cell = summary_sheet.cell(row=1, column=status_col_idx)
                status_cell.value = 'Most Adcalls status'
                status_cell.fill = header_fill
                status_cell.font = header_font
                status_cell.alignment = header_alignment
                status_cell.border = header_border
                
                pct_cell = summary_sheet.cell(row=1, column=pct_col_idx)
                pct_cell.value = 'Most Ad Calls %'
                pct_cell.fill = header_fill
                pct_cell.font = header_font
                pct_cell.alignment = header_alignment
                pct_cell.border = header_border
                
                # Set column widths
                summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(status_col_idx)].width = 25
                summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(pct_col_idx)].width = 15
                
                # Debug: Print all error types in the data
                print("\nDEBUG: Error types in the data:")
                for website, data in error_dist_summary.items():
                    print(f"Website: {website}, Type: {data.get('Type', 'Not found')}, Status: {data.get('Most Adcalls status', 'Not found')}")
                
                # Add error distribution summary data for each website
                for row_idx in range(2, summary_sheet.max_row + 1):
                    website_cell = summary_sheet.cell(row=row_idx, column=website_col_idx)
                    website_name = website_cell.value
                    
                    if website_name in error_dist_summary:
                        # Add Most Adcalls status
                        status_value = error_dist_summary[website_name]['Most Adcalls status']
                        status_cell = summary_sheet.cell(row=row_idx, column=status_col_idx)
                        status_cell.value = status_value
                        
                        # Add Most Ad Calls %
                        pct_value = error_dist_summary[website_name]['Most Ad Calls %']
                        pct_cell = summary_sheet.cell(row=row_idx, column=pct_col_idx)
                        pct_cell.value = pct_value
                        
                        # Get the error type and apply formatting right away
                        error_type = error_dist_summary[website_name].get('Type', 'UNKNOWN')
                        print(f"Applying formatting for {website_name}: Type={error_type}")
                        
                        # Apply conditional formatting based on Type value
                        # Case-insensitive comparison to be more robust
                        error_type_upper = str(error_type).upper()
                        
                        if 'OK' in error_type_upper:
                            # Light green with black letters
                            status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            status_cell.font = Font(color='000000')
                        elif 'WARNING_FIXED' in error_type_upper:
                            # Very light yellow with black letters
                            status_cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                            status_cell.font = Font(color='000000')
                        elif 'WARNING' in error_type_upper:
                            # Orange with white letters
                            status_cell.fill = PatternFill(start_color='FFB266', end_color='FFB266', fill_type='solid')
                            status_cell.font = Font(color='FFFFFF')
                        elif 'ERROR' in error_type_upper:
                            # Red with white letters
                            status_cell.fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
                            status_cell.font = Font(color='FFFFFF')
                        else:
                            # Light blue for Unknown
                            status_cell.fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
                            status_cell.font = Font(color='000000')
        
        # # LEGENDS TEMPORARILY DISABLED
        # # Add legends for conditional formatting in the Summary sheet
        # if 'Summary' in book.sheetnames:
        #     summary_sheet = book['Summary']
        #     
        #     # Get column P index for Most Adcalls status legend
        #     p_col_idx = None
        #     for col_idx, cell in enumerate(summary_sheet[1], 1):
        #         if cell.value == 'Most Adcalls status':
        #             p_col_idx = col_idx
        #             break
        #     
        #     # If not found, use column P
        #     if not p_col_idx:
        #         p_col_idx = openpyxl.utils.column_index_from_string('P')
        #     
        #     # Define legend for Most Adcalls status (Column P)
        #     most_adcalls_legend_items = [
        #         {'label': 'OK', 'bg_color': '00635D', 'font_color': 'FFFFFF'},
        #         {'label': 'WARNING_FIXED', 'bg_color': 'BEE7B8', 'font_color': '000000'},
        #         {'label': 'WARNING', 'bg_color': 'BE7C4D', 'font_color': 'FFFFFF'},
        #         {'label': 'ERROR', 'bg_color': '4C212A', 'font_color': 'FFFFFF'},
        #         {'label': 'UNKNOWN', 'bg_color': '01172F', 'font_color': 'FFFFFF'}
        #     ]
        #     
        #     # Add legend for Most Adcalls status at position P1
        #     p_legend_pos = add_formatting_legends(summary_sheet, 
        #                                      'Conditional Formatting Legend for Most Ad Calls', 
        #                                      most_adcalls_legend_items,
        #                                      start_cell=(1, p_col_idx),
        #                                      include_descriptions=False)
        #     
        #     # Define legend for Status (Column I)
        #     status_legend_items = [
        #         {'label': 'Monetisation OK', 'bg_color': '00635D', 'font_color': '000000', 'description': '300_Good'},
        #         {'label': 'Monetisation blocked with at least one Bidder', 'bg_color': 'BE7C4D', 'font_color': '000000', 'description': '200_Warning'},
        #         {'label': 'Monetisation blocked for this domain', 'bg_color': '4C212A', 'font_color': 'FFFFFF', 'description': '100_error'},
        #         {'label': 'Site inactive or Archived', 'bg_color': '6F70A1', 'font_color': 'FFFFFF', 'description': '500_inactive 700_archived'}
        #     ]
        #     
        #     # Add legend for Status below the first legend
        #     next_legend_row = p_legend_pos[0] + 2  # Add a bit more space
        #     status_legend_pos = add_formatting_legends(summary_sheet, 
        #                                      'Status', 
        #                                      status_legend_items,
        #                                      start_cell=(next_legend_row, p_col_idx))
        #     
        #     # Define legend for metrics percentage (used in various columns)
        #     metrics_legend_items = [
        #         {'label': 'High', 'bg_color': 'FFCCCC', 'font_color': '000000', 'description': 'High value (>80%)'},
        #         {'label': 'Medium', 'bg_color': 'FFFFCC', 'font_color': '000000', 'description': 'Medium value (50-80%)'},
        #         {'label': 'Low', 'bg_color': 'CCFFCC', 'font_color': '000000', 'description': 'Low value (20-50%)'},
        #         {'label': 'Very Low', 'bg_color': 'FFFFFF', 'font_color': '000000', 'description': 'Very low value (<20%)'}
        #     ]
        #     
        #     # Add legend for metrics below the status legend, but 3 columns to the left
        #     metrics_column = max(1, p_col_idx - 3)  # Ensure we don't go below column 1
        #     add_formatting_legends(summary_sheet, 
        #                          'Metrics Percentage', 
        #                          metrics_legend_items,
        #                          start_cell=(status_legend_pos[0] + 2, metrics_column))
        
        # Save the workbook
        book.save(output_path)
        
        debug_print("Error distribution data added successfully")
        return True
    except Exception as e:
        debug_print(f"Error writing error distribution data: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def apply_consistent_formatting(excel_path, domains_df, rules, error_dist_summary=None):
    """
    Apply consistent conditional formatting to all sheets in the workbook.
    This is a separate pass to ensure formatting is identical across sheets.
    
    Parameters:
        excel_path (str): Path to the Excel file
        domains_df (DataFrame): DataFrame containing domains highlight data
        rules (dict): Dictionary of rules to apply
    """
    try:
        print(f"Phase 2: Applying consistent formatting to all sheets in {excel_path}")
        
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)
        
        # Use styles from config
        styles = FORMATTING_CONFIG['custom_styles']
        
        # Process each domain in the domains DataFrame
        domain_priorities = {}
        
        # First pass: Evaluate all domains against rules to determine priorities
        print("\nEvaluating rules for each domain:")
        
        # Print debug info about the rules
        for priority, priority_rules in rules.items():
            print(f"\n{priority} Priority Rules:")
            for i, rule in enumerate(priority_rules):
                print(f"  Rule {i+1}: {rule.get('metric', '')} {rule.get('operator', '')} {rule.get('value', '')}")
        
        for row_idx, row in domains_df.iterrows():
            domain_name = row.get('Website/App Name', '')
            if not domain_name:
                continue
                
            # Evaluate rules for this domain
            for priority, priority_rules in rules.items():
                all_match = True
                rule_debug_info = []
                
                for rule in priority_rules:
                    # Check if all rules for this priority match
                    metric = rule.get('metric')
                    operator = rule.get('operator')
                    value = rule.get('value')
                    
                    if not all([metric, operator, value]) or metric not in row:
                        all_match = False
                        rule_debug_info.append(f"{metric} {operator} {value} => SKIPPED (metric not found)")
                        break
                        
                    # Evaluate the rule
                    cell_value = row[metric]
                    rule_applies = evaluate_rule(cell_value, operator, value)
                    
                    # For green highlight, add special debug output
                    if priority == "Low" and operator == '>' and float(value) == 5:
                        print(f"Checking Green Highlight for {domain_name}: {metric}={cell_value} {operator} {value} => {rule_applies}")
                    
                    rule_debug_info.append(f"{metric}={cell_value} {operator} {value} => {rule_applies}")
                    
                    if not rule_applies:
                        all_match = False
                        break
                
                # Debug output for rules that matched or failed
                if all_match:
                    print(f"Domain {domain_name} matches {priority} priority: {'; '.join(rule_debug_info)}")
                    domain_priorities[domain_name] = priority
                    break  # Use highest priority that applies
        
        print(f"Determined priorities for {len(domain_priorities)} domains")
        
        # Now apply consistent formatting to all sheets based on domain name
        sheets_to_format = ['Domains Highlight']
        
        # Add Summary sheet if configured
        if FORMATTING_CONFIG['format_summary_sheet'] and 'Summary' in workbook.sheetnames:
            sheets_to_format.append('Summary')
            
        # Add Supply Chain sheet if configured
        if FORMATTING_CONFIG['format_supply_chain'] and 'Supply Chain Validation' in workbook.sheetnames:
            sheets_to_format.append('Supply Chain Validation')
        
        for sheet_name in sheets_to_format:
            if sheet_name not in workbook.sheetnames:
                print(f"Sheet {sheet_name} not found, skipping")
                continue
                
            sheet = workbook[sheet_name]
            
            # Find the domain column
            domain_col_idx = None
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value == 'Website/App Name':
                    domain_col_idx = col_idx
                    break
                    
            if domain_col_idx is None:
                print(f"Domain column not found in {sheet_name}, skipping")
                continue
                
            # No special handling needed for columns
            
            # Find Status column (in Summary sheet)
            status_col_idx = None
            info_col_idx = None
            if sheet_name == 'Summary':
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value == 'Status':
                        status_col_idx = col_idx
                    # Column J will be one column after Status column I
                    if cell.value == 'Missing Primary Bidders':
                        info_col_idx = col_idx
            
            # Apply formatting based on domain priorities
            for row_idx in range(2, sheet.max_row + 1):
                domain_cell = sheet.cell(row=row_idx, column=domain_col_idx)
                domain_name = domain_cell.value
                
                # Apply Status column conditional formatting
                if sheet_name == 'Summary' and status_col_idx is not None:
                    status_cell = sheet.cell(row=row_idx, column=status_col_idx)
                    
                    # Add conditional formatting for Status column
                    if status_cell.value == '250_info':
                        # Light blue formatting for 250_info
                        status_cell.fill = PatternFill(
                            start_color='ADD8E6',  # Light blue
                            end_color='ADD8E6',
                            fill_type='solid'
                        )
                        
                        # Add "missing secondary lines" in the next column (column J)
                        if info_col_idx is not None:
                            info_cell = sheet.cell(row=row_idx, column=info_col_idx)
                            info_cell.value = 'missing secondary lines'
                            
                    elif status_cell.value == '300_bypassed':
                        # Light orange formatting for 300_bypass
                        status_cell.fill = PatternFill(
                            start_color='FFD8B1',  # Light orange
                            end_color='FFD8B1',
                            fill_type='solid'
                        )
                
                if not domain_name or domain_name not in domain_priorities:
                    continue
                    
                priority = domain_priorities[domain_name]
                style = styles.get(priority)
                
                if style:
                    print(f"Applying {priority} formatting to {domain_name} in {sheet_name}")
                    
                    # Debug - print rule evaluation details for green highlight filter
                    if priority == "Low":
                        print(f"Green highlight (Low priority) applied to {domain_name} in {sheet_name}")
                    
                    if FORMATTING_CONFIG['apply_to_entire_row']:
                        # Apply to entire row
                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            
                            # No special handling needed for any columns
                                
                            cell.fill = PatternFill(
                                start_color=style['bg_color'], 
                                end_color=style['bg_color'], 
                                fill_type='solid'
                            )
                            
                            # Only change font color for the domain cell
                            if col_idx == domain_col_idx:
                                cell.font = Font(color=style['font_color'])
                    else:
                        # Apply formatting to domain cell only
                        domain_cell.fill = PatternFill(
                            start_color=style['bg_color'], 
                            end_color=style['bg_color'], 
                            fill_type='solid'
                        )
                        domain_cell.font = Font(color=style['font_color'])
        
        # Save the workbook
        workbook.save(excel_path)
        print(f"Formatting applied successfully to {excel_path}")
        return True
        
    except Exception as e:
        print(f"Error applying formatting: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def create_summary_sheet(output_path, domains_df, supply_chain_df, apply_formatting=True):
    """
    Create a summary sheet that combines key metrics from domains highlight and supply chain validation.
    Performs a lookup between Sheet 1 and Sheet 2 and displays the combined data.
    
    Parameters:
        output_path (str): Path to the Excel file
        domains_df (DataFrame): DataFrame containing domains highlight data
        supply_chain_df (DataFrame): DataFrame containing supply chain validation data
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print("\n" + "=" * 80)
        print("SUMMARY SHEET CREATION")
        print("=" * 80)
        
        print("\nDOMAINS DataFrame information:")
        print(f"- Shape: {domains_df.shape}")
        print(f"- Columns: {domains_df.columns.tolist()}")
        print("\nSample rows from DOMAINS DataFrame:")
        print(domains_df.head(3).to_string())
        
        print("\nSUPPLY CHAIN DataFrame information:")
        print(f"- Shape: {supply_chain_df.shape}")
        print(f"- Columns: {supply_chain_df.columns.tolist()}")
        print("\nSample rows from SUPPLY CHAIN DataFrame:")
        print(supply_chain_df.head(3).to_string())
        
        print("\n" + "=" * 80)
        
        debug_print(f"Creating summary sheet in {output_path}")
        
        # Print supply chain DataFrame column names for debugging
        debug_print(f"Supply Chain columns: {supply_chain_df.columns.tolist()}")
        
        # Print a sample row from supply chain data
        if not supply_chain_df.empty:
            debug_print(f"Supply Chain sample row: {supply_chain_df.iloc[0].to_dict()}")
        
        # Use Website/App Name if available
        domain_col = None
        for col in ['Website Domain', 'Website/App Name', 'Website Name', 'Domain']:
            if col in supply_chain_df.columns:
                domain_col = col
                debug_print(f"Using {col} as domain column from supply chain data")
                break
                
        if not domain_col:
            debug_print("Warning: No domain column found in supply chain data")
            return False
        
        # Print the actual column names from both dataframes
        print("\nDomains Highlight column names:")
        for i, col in enumerate(domains_df.columns):
            print(f"  Column {i}: {col}")
            
        print("\nSupply Chain column names:")
        for i, col in enumerate(supply_chain_df.columns):
            print(f"  Column {i}: {col}")
        
        # We need to match Website/App Name from domains_df with Name from supply_chain_df
        domain_col = 'Website/App Name'
        name_col = 'Name'  # The matching column in supply chain data
        
        # Verify these columns exist
        if domain_col not in domains_df.columns:
            print(f"ERROR: Column '{domain_col}' not found in Domains Highlight data")
            return False
            
        if name_col not in supply_chain_df.columns:
            print(f"ERROR: Column '{name_col}' not found in Supply Chain data")
            return False
            
        # Identify missing lines columns in supply chain data
        primary_col = 'Primary Missing' if 'Primary Missing' in supply_chain_df.columns else None
        secondary_col = 'Secondary Missing' if 'Secondary Missing' in supply_chain_df.columns else None
        status_col = 'Status' if 'Status' in supply_chain_df.columns else None
        
        print(f"\nUsing the following columns for matching:")
        print(f"- Domains Highlight: '{domain_col}'")
        print(f"- Supply Chain: '{name_col}'")
        print(f"- Supply Chain (missing lines): '{primary_col}', '{secondary_col}', '{status_col}'")
        
        # Create a lookup dictionary from the supply chain data
        # Use the Name column as the key for matching with Website/App Name
        lookup_dict = {}
        
        # Print a few sample rows from supply chain data to debug
        print("\nSample data from Supply Chain:")
        sample_rows = min(5, len(supply_chain_df))
        for i in range(sample_rows):
            row = supply_chain_df.iloc[i]
            name = row.get(name_col, 'N/A')
            primary = row.get(primary_col, 'N/A') if primary_col else 'N/A'
            secondary = row.get(secondary_col, 'N/A') if secondary_col else 'N/A'
            status = row.get(status_col, 'N/A') if status_col else 'N/A'
            print(f"  Row {i}: Name='{name}', Primary={primary}, Secondary={secondary}, Status={status}")
        
        # Add each name and its data to the lookup dictionary
        for _, row in supply_chain_df.iterrows():
            name_value = row.get(name_col, '')
            
            # Skip rows with empty names
            if pd.isna(name_value) or name_value == '':
                continue
            
            # Normalize name (lowercase, strip whitespace)
            name = str(name_value).lower().strip()
            
            # Get the missing lines counts and status
            primary_missing = 0
            secondary_missing = 0
            status = ''
            
            if primary_col and primary_col in row:
                primary_value = row[primary_col]
                if not pd.isna(primary_value):
                    try:
                        primary_missing = int(primary_value)
                    except (ValueError, TypeError):
                        print(f"Warning: Could not convert '{primary_value}' to int for name '{name}'")
            
            if secondary_col and secondary_col in row:
                secondary_value = row[secondary_col]
                if not pd.isna(secondary_value):
                    try:
                        secondary_missing = int(secondary_value)
                    except (ValueError, TypeError):
                        print(f"Warning: Could not convert '{secondary_value}' to int for name '{name}'")
            
            if status_col and status_col in row:
                status_value = row[status_col]
                if not pd.isna(status_value):
                    status = str(status_value)
            
            # Check for Missing Primary Bidders column
            missing_bidders = ''
            if 'Missing Primary Bidders' in row:
                missing_bidders_value = row['Missing Primary Bidders']
                if not pd.isna(missing_bidders_value):
                    missing_bidders = str(missing_bidders_value)
                    
            # Store in lookup dictionary with name as key
            lookup_dict[name] = {
                'Primary Missing': primary_missing,
                'Secondary Missing': secondary_missing,
                'Status': status,
                'Missing Primary Bidders': missing_bidders
            }
        
        print(f"Created lookup dictionary with {len(lookup_dict)} entries from Supply Chain Names")
        
        debug_print(f"Created lookup dictionary with {len(lookup_dict)} entries")
        
        # Create a new dataframe for the summary sheet
        summary_columns = [
            'Website/App Name',
            'Ad Requests', 
            'Revenue', 
            'RPB', 
            'Bid Rate', 
            'Win Rate',
            'Primary Missing',
            'Secondary Missing',
            'Status',
            'Missing Primary Bidders'
        ]
        
        summary_df = pd.DataFrame(columns=summary_columns)
        
        # Create a direct mapping from domains_df to supply_chain_df
        # Matching Website/App Name from domains_df with Name from supply_chain_df
        print("\nMatching Website/App Name from Domains Highlight with Name from Supply Chain...")
        
        # Fill in data from domains df
        match_count = 0
        no_match_count = 0
        
        # No need to track domain to websites mapping anymore
        
        rows_list = []  # Use a list to store rows instead of concat for better performance
        
        for _, row in domains_df.iterrows():
            website_name = row.get(domain_col, '')
            
            # Skip non-string or empty website names
            if not isinstance(website_name, str) or not website_name:
                continue
                
            # Normalize website name (lowercase, strip whitespace)
            website_name_norm = website_name.lower().strip()
            
            # Create a new row with domains data
            new_row = {
                'Website/App Name': website_name,  # Keep original capitalization
                'Ad Requests': row.get('Ad Requests', np.nan),
                'Revenue': row.get('Revenue', np.nan),
                'RPB': row.get('RPB', np.nan),
                'Bid Rate': row.get('Bid Rate', np.nan),
                'Win Rate': row.get('Win Rate', np.nan),
                'Primary Missing': 0,  # Default to 0 instead of NaN
                'Secondary Missing': 0,  # Default to 0 instead of NaN
                'Status': '',
                'Missing Primary Bidders': ''
            }
            
            # Check if we have supply chain data for this website name
            found_match = False
            
            # Try exact match first
            if website_name_norm in lookup_dict:
                supply_data = lookup_dict[website_name_norm]
                new_row['Primary Missing'] = supply_data.get('Primary Missing', 0)
                new_row['Secondary Missing'] = supply_data.get('Secondary Missing', 0) 
                new_row['Status'] = supply_data.get('Status', '')
                new_row['Missing Primary Bidders'] = supply_data.get('Missing Primary Bidders', '')
                print(f"Match for '{website_name}': Primary={new_row['Primary Missing']}, Secondary={new_row['Secondary Missing']}, Status={new_row['Status']}")
                found_match = True
                match_count += 1
            
            # No match found - try a fuzzy match on each part of the name
            if not found_match:
                # Try matching any part of the name
                for lookup_name in lookup_dict.keys():
                    # Check if lookup_name is part of website_name or vice versa
                    if (lookup_name in website_name_norm or 
                        website_name_norm in lookup_name or
                        any(part in lookup_name for part in website_name_norm.split()) or
                        any(part in website_name_norm for part in lookup_name.split())):
                        
                        supply_data = lookup_dict[lookup_name]
                        new_row['Primary Missing'] = supply_data.get('Primary Missing', 0)
                        new_row['Secondary Missing'] = supply_data.get('Secondary Missing', 0)
                        new_row['Status'] = supply_data.get('Status', '')
                        new_row['Missing Primary Bidders'] = supply_data.get('Missing Primary Bidders', '')
                        print(f"Fuzzy match: '{website_name}' ⟷ '{lookup_name}': Primary={new_row['Primary Missing']}, Secondary={new_row['Secondary Missing']}, Status={new_row['Status']}")
                        found_match = True
                        match_count += 1
                        break
            
            # No match found at all
            if not found_match:
                no_match_count += 1
                
            # Add the row to our list
            rows_list.append(new_row)
        
        # Create the summary DataFrame from our list of rows
        summary_df = pd.DataFrame(rows_list)
        
        print(f"Matching complete: {match_count} matches, {no_match_count} website names without matches")
        
        # Sort by Revenue (descending)
        if 'Revenue' in summary_df.columns:
            summary_df = summary_df.sort_values('Revenue', ascending=False)
        
        # Write to Excel
        debug_print(f"Writing summary sheet with {len(summary_df)} rows")
        
        # Load the workbook
        book = openpyxl.load_workbook(output_path)
        
        # Check if Summary sheet already exists
        sheet_name = 'Summary'
        if sheet_name in book.sheetnames:
            debug_print(f"{sheet_name} already exists, removing it first")
            idx = book.sheetnames.index(sheet_name)
            book.remove(book.worksheets[idx])
        
        # Save the workbook
        book.save(output_path)
        
        # Write the summary data
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting to the summary sheet
        book = openpyxl.load_workbook(output_path)
        sheet = book[sheet_name]
        
        # Format headers
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(bottom=Side(style='medium', color='000000'))
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
            
        # We now handle error distribution column formatting directly in write_error_distribution_sheet
        # So no additional formatting is needed here for the error distribution columns
            
        # Skip formatting in this function when apply_formatting is False
        # The formatting will be handled by apply_consistent_formatting instead
        if not apply_formatting:
            print("Skipping formatting in create_summary_sheet - will be handled by separate formatting pass")
        
        # Apply basic numeric formatting without conditional formatting
        # Format numeric columns
        for row_idx in range(2, len(summary_df) + 2):  # Skip header
            # Format Revenue as currency
            revenue_cell = sheet.cell(row=row_idx, column=summary_columns.index('Revenue') + 1)
            if revenue_cell.value is not None and isinstance(revenue_cell.value, (int, float)):
                revenue_cell.number_format = '$#,##0.00'
            
            # Format percentage columns
            for col_name in ['Bid Rate', 'Win Rate']:
                col_idx = summary_columns.index(col_name) + 1
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            column_header = sheet.cell(row=1, column=column[0].column).value
            
            # Set specific width for Related Websites column
            if column_header == 'Related Websites':
                sheet.column_dimensions[column_letter].width = 60  # Fixed larger width for related websites
                # Also enable text wrapping for better readability
                for cell in column:
                    if cell.row > 1:  # Skip header
                        cell.alignment = Alignment(wrap_text=True)
            else:
                # For other columns, calculate based on content
                for cell in column:
                    if cell.row == 1:  # Header row
                        max_length = max(max_length, len(str(cell.value)) + 2)  # Extra padding for headers
                    else:
                        max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format numeric columns
        for row_idx in range(2, len(summary_df) + 2):  # Skip header
            # Format Revenue as currency
            revenue_cell = sheet.cell(row=row_idx, column=summary_columns.index('Revenue') + 1)
            if revenue_cell.value is not None and isinstance(revenue_cell.value, (int, float)):
                revenue_cell.number_format = '$#,##0.00'
            
            # Format percentage columns
            for col_name in ['Bid Rate', 'Win Rate']:
                col_idx = summary_columns.index(col_name) + 1
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00%'
            
            # Color the status column based on value
            status_idx = summary_columns.index('Status') + 1
            status_cell = sheet.cell(row=row_idx, column=status_idx)
            
            if status_cell.value:
                status_value = str(status_cell.value).lower()
                if 'error' in status_value or 'critical' in status_value or '100_error' in status_value:
                    status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
                    status_cell.font = Font(color='FFFFFF')
                elif 'warning' in status_value or '200_warning' in status_value:
                    status_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                elif 'ok' in status_value or 'good' in status_value or '200' in status_value or '400_valid' in status_value:
                    status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light Green
                elif '500_inactive' in status_value or '700_archived' in status_value:
                    status_cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Light Grey
                    status_cell.font = Font(color='000000')
        
        # Save the workbook
        book.save(output_path)
        
        debug_print("Summary sheet created successfully")
        return True
    except Exception as e:
        debug_print(f"Error creating summary sheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def apply_highlighting_rules(excel_path, df, rules, sheet_name='Domains Highlight', apply_formatting=True):
    """
    Apply highlighting rules to the Excel output
    
    Parameters:
        excel_path (str): Path to the Excel file
        df (DataFrame): DataFrame containing the data
        rules (dict): Dictionary of rules to apply
        sheet_name (str): Name of the sheet to apply rules to
    """
    debug_print(f"Starting to apply highlighting rules to {excel_path}, sheet={sheet_name}")
    debug_print(f"DataFrame columns: {df.columns.tolist()}")
    
    # Print the rules in a more readable format
    debug_print("Rules configuration:")
    for priority, priority_rules in rules.items():
        debug_print(f"  {priority} Priority:")
        for rule in priority_rules:
            debug_print(f"    {rule['metric']} {rule['operator']} {rule['value']}")
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_path)
        
        # Get the worksheet
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        
        # Define styles for each priority level
        styles = {
            "High": {
                "bg_color": "FFC7CE",  # Light red
                "font_color": "9C0006"  # Dark red
            },
            "Medium": {
                "bg_color": "FFEB9C",  # Light yellow
                "font_color": "9C6500"  # Dark yellow
            },
            "Low": {
                "bg_color": "C6EFCE",  # Light green
                "font_color": "006100"  # Dark green
            }
        }
        
        # Track which rows match each priority level
        row_highlighting = {}
        
        debug_print(f"Processing {len(df)} rows for highlighting")
        
        # Check each row against the rules
        for row_idx, row in df.iterrows():
            # Start with highest priority and work down
            for priority in ["High", "Medium", "Low"]:
                if priority not in rules:
                    continue
                    
                priority_rules = rules.get(priority, [])
                if not priority_rules:
                    continue
                    
                # Check if all rules for this priority level match
                all_match = True
                for rule in priority_rules:
                    metric = rule.get("metric")
                    operator = rule.get("operator")
                    value = rule.get("value")
                    
                    # Get the column name for this metric
                    cell_value = None
                    column_used = None
                    
                    # First, try to find an exact match for the metric name
                    if metric in df.columns:
                        cell_value = row[metric]
                        column_used = metric
                        debug_print(f"Rule ({priority}) - Found exact match for {metric}, value={cell_value}")
                    else:
                        # If no exact match, look for it in column names
                        for col in df.columns:
                            # Look for the metric name in the column
                            if metric in col:
                                # Skip % Diff columns unless we're specifically looking for percentage differences
                                if "% Diff" in col:
                                    if metric.endswith(" % Diff"):
                                        cell_value = row[col]
                                        column_used = col
                                        debug_print(f"Rule ({priority}) - Found diff match for {metric}, column={col}, value={cell_value}")
                                        break
                                else:
                                    # For base metrics like Revenue, RPB, etc.
                                    cell_value = row[col]
                                    column_used = col
                                    debug_print(f"Rule ({priority}) - Found partial match for {metric}, column={col}, value={cell_value}")
                                    break
                            
                    if cell_value is None or pd.isna(cell_value):
                        # This metric doesn't exist in the data
                        all_match = False
                        break
                        
                    # Process cell value and rule value
                    try:
                        # Convert cell value to float if it's not a string
                        if not isinstance(cell_value, str):
                            cell_value_float = float(cell_value)
                        # For strings, handle percentage values
                        elif isinstance(cell_value, str) and cell_value.endswith('%'):
                            cell_value_float = float(cell_value.rstrip('%')) / 100
                        else:
                            cell_value_float = float(cell_value)
                            
                        debug_print(f"Converted cell value {cell_value} to {cell_value_float}")
                    except ValueError as e:
                        debug_print(f"Error converting cell value to float: {str(e)}")
                        all_match = False
                        break
                        
                    # Handle rule value based on operator
                    if operator == "Between":
                        # For Between, keep value as a string until later
                        rule_value = value
                        debug_print(f"Between operator - keeping rule value as string: {rule_value}")
                    else:
                        # For Rate metrics, handle percentage values intelligently
                        if "Rate" in metric:
                            if isinstance(value, str) and value.endswith('%'):
                                rule_value = float(value.rstrip('%')) / 100
                            else:
                                # Convert to float first
                                rule_value = float(value)
                        else:
                            # For non-percentage values, just convert to float
                            rule_value = float(value)
                            
                        # Check if we need to convert percentage cell values
                        if isinstance(cell_value, str) and '%' in cell_value:
                            try:
                                # Remove % and convert to decimal
                                cell_value = float(cell_value.replace('%', '')) / 100
                            except ValueError:
                                debug_print(f"Could not convert percentage: {cell_value}")
                                all_match = False
                                break
                    
                    # Skip if cell_value is missing
                    if pd.isna(cell_value):
                        debug_print(f"Skipping rule {metric} {operator} {value} due to missing value")
                        all_match = False
                        break
                    
                    # Use our evaluate_rule helper function
                    rule_applies = evaluate_rule(cell_value, operator, value)
                    
                    debug_print(f"Rule ({priority}) - {metric} {operator} {value} - Applied to {column_used}={cell_value} - Result: {rule_applies}")
                            
                    if not rule_applies:
                        all_match = False
                
                        
                if all_match:
                    row_highlighting[row_idx] = priority
                    domain_name = df.iloc[row_idx]['Website/App Name'] if 'Website/App Name' in df.columns else f"Row {row_idx}"
                    debug_print(f"Row {row_idx} ({domain_name}) highlighted as {priority} priority")
                    break  # Only apply the highest priority satisfied
        
        # Apply highlighting to rows
        debug_print(f"Applying highlighting to {len(row_highlighting)} rows")
        for row_idx, priority in row_highlighting.items():
            # Get Excel row index (add 2 for header and 1-based indexing)
            excel_row = row_idx + 2
            style = styles[priority]
            
            # Find the domain/website name for better debugging
            domain_name = df.iloc[row_idx]['Website/App Name'] if 'Website/App Name' in df.columns else f"Row {row_idx}"
            
            # Apply to all cells in the row for better visibility
            debug_print(f"Applying {priority} formatting to row {excel_row} ({domain_name})")
            
            # Apply to the entire row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.fill = PatternFill(start_color=style['bg_color'], end_color=style['bg_color'], fill_type='solid')
                
                # Only change font color for the first column (domain name)
                if col_idx == 1:
                    cell.font = Font(color=style['font_color'])

        if apply_formatting:
            # Format percentage difference columns
            for col_idx, col_name in enumerate(df.columns):
                if '% Diff' in col_name:
                    # Apply percentage format
                    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    for row in range(2, len(df) + 2):  # Skip header row
                        cell = worksheet.cell(row=row, column=col_idx + 1)
                        cell.number_format = '0.00%'
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        # Save the workbook
        workbook.save(excel_path)
        return True
            
    except Exception as e:
        print(f"Error applying highlighting: {str(e)}")
        return False
